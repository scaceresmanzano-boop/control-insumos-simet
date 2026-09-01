"""Carga inicial (idempotente) del catálogo de insumos desde el Excel consolidado de la tesis.

Se apoya en posiciones de columna fijas (no en texto de encabezado) porque el Excel
tiene encabezados fuera de la fila 1 y texto con acentos/saltos de línea que conviene
no usar como clave de matching.
"""
from pathlib import Path
import openpyxl

import db

BASE_DIR = Path(__file__).parent
ARCHIVO_CONSOLIDADO = BASE_DIR / "Control_Insumos_Consolidado_SIMET-USACH.xlsx"

HOJA_CATALOGO = "Catálogo de Insumos"
FILA_INICIO = 12  # fila 11 es el ejemplo (ID 0), se excluye

# Índice de columna (1-based) -> nombre de campo en la tabla insumos.
COLUMNAS = {
    1: "id",
    2: "nombre",
    3: "categoria",
    4: "operacion",
    5: "costo_unitario",
    6: "unidades_por_cambio",
    8: "frecuencia_uso_mensual",
    9: "rendimiento_probetas",
    11: "impacto_operacional",
    12: "tiempo_reposicion_dias",
    24: "fuente_dato",
    25: "ubicacion",
}


def _seed_insumos(conn):
    wb = openpyxl.load_workbook(ARCHIVO_CONSOLIDADO, data_only=True)
    ws = wb[HOJA_CATALOGO]
    filas = []
    for row in ws.iter_rows(min_row=FILA_INICIO, max_row=ws.max_row):
        valores = {campo: row[col - 1].value for col, campo in COLUMNAS.items()}
        if valores["id"] is None or not valores["nombre"]:
            continue
        valores["id"] = int(valores["id"])
        filas.append(valores)

    campos = list(COLUMNAS.values())
    for v in filas:
        set_clause = ", ".join(f"{c} = excluded.{c}" for c in campos if c != "id")
        conn.execute(
            f"""
            INSERT INTO insumos ({', '.join(campos)})
            VALUES ({', '.join('?' for _ in campos)})
            ON CONFLICT(id) DO UPDATE SET {set_clause}
            """,
            [v[c] for c in campos],
        )
    conn.commit()
    return len(filas)


def seed():
    conn = db.get_connection()
    db.init_schema(conn)
    n_insumos = _seed_insumos(conn)
    db.asignar_codigos_automaticos(conn)
    conn.close()
    return n_insumos


if __name__ == "__main__":
    n_insumos = seed()
    print(f"Insumos cargados/actualizados: {n_insumos}")
